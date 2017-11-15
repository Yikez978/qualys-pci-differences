#!/usr/bin/env python
### Compile metrics to report count of discovered vulns and remediated vulns from Qualys scans
### November 9, 2017
### Melissa Canazon


import argparse
import openpyxl
from openpyxl import load_workbook


parser = argparse.ArgumentParser(prog='boardmetrics')
parser.add_argument('-D', '--Debug', help='Debug Mode assists in determining issues being raised by the script.', action='store_true')
parser.add_argument('-L', '--lastScan', help='last months scan')
parser.add_argument('-N', '--newScan', help='new months scan')
parser.add_argument('-M', '--metricsFile', help='name of new metrics file')
parser.add_argument('-R', '--reportDate', help='month and year of the report')
args = parser.parse_args()


# last months results
lastS = load_workbook(args.lastScan, data_only=True)
lastResults = lastS.get_sheet_by_name(lastS.get_sheet_names()[0])
maxrL = str(lastResults.max_row -2)

# current months results
newS = load_workbook(args.newScan, data_only=True)
newResults = newS.get_sheet_by_name(newS.get_sheet_names()[0])
maxrN = str(newResults.max_row -2)

# Vuln titles not counted
disabledTitles = ['Possible Scan Interference', 'Web Server Stopped Responding', 'Exhaustive Web Testing Skipped', 'Service Stopped Responding']

#create dictionary of vuln and count for last month
#dictionary by severity with a value of a dictionary of each vuln and count
#{'low':{'NTP Information Disclosure Vulnerability': 33, 'Vuln title': 26}, 'medium' : {'vuln':7}, 'high': {}, 'critical':{}}

def vulnDict(results, maxrow):

    dcount = {'low' : {}, 'medium' : {}, 'high' : {}, 'critical' : {}}
    results = results
    maxrow = maxrow

    for rows in results['A9' : 'A' + maxrow]:
        for cell in rows:
            row = str(cell.row)
            if results['Q' + row].value != None:
                if results['G' + row].value not in disabledTitles:
                    title = results['G' + row].value
                    cvssBase = float((results['Q' + row].value).split(' ')[0])

                    if args.Debug:
                        print('cvss base: ' + str(cvssBase))

                    if cvssBase < 4:
                        severity = 'low'
                    elif 4 <= cvssBase < 7:
                        severity = 'medium'
                    elif 7 <= cvssBase < 10:
                        severity = 'high'
                    elif cvssBase >= 10:
                        severity = 'critical'
                    else:
                        print('cvssBase: ' + str(cvssBase))

                    if args.Debug:
                        print(severity)

                    if title in dcount[severity]:
                        count = dcount[severity][title]
                        dcount[severity][title] = count + 1
                        if args.Debug:
                            print(title + ': count = ' + str(dcount[severity][title]) + '\n')
                    else:
                        dcount[severity][title] = 1
                        if args.Debug:
                            print(title + ': count = ' + str(dcount[severity][title]) + '\n')

    return dcount

#total of each severity

def totalSeverity(dcount):
    dseverity = {'low' : 0, 'medium' : 0, 'high' : 0, 'critical' :0}
    for severity in dcount:
        if args.Debug:
            print(severity)

        for vuln in dcount[severity]:
            # print(str(dcount[severity][vuln]))
            dseverity[severity] += dcount[severity][vuln]
            if args.Debug:
                print(str(dseverity[severity]))
        print(severity + ': ' + str(dseverity[severity]))
    return dseverity


#sum of differences between vulns (newly discoverd vs. remediated)
    # difference = new - old

def differences(dlast, dnew):
    difference = {'low' : {'discovered' : 0, 'remediated' : 0}, 'medium' : {'discovered' : 0, 'remediated' : 0}, 'high'  : {'discovered' : 0, 'remediated' : 0}, 'critical'  : {'discovered' : 0, 'remediated' : 0}}

    for severity in dlast:
        for vuln in dlast[severity]:
            if vuln in dnew[severity]:
                diff = dnew[severity][vuln] - dlast[severity][vuln]
            else:
                diff = 0 - dlast[severity][vuln]

            if diff <= 0:
                difference[severity]['remediated'] += diff
            else:
                difference[severity]['discovered'] += diff


    for severity in dnew:
        for vuln in dnew[severity]:
            if vuln in dlast[severity]:
                pass
            else:
                difference[severity]['discovered'] += dnew[severity][vuln]

    if args.Debug:
        print(str(difference))

    return difference

# MAIN
dlastCounts = vulnDict(lastResults, maxrL)
print('Last Counts Returned')

dnewCounts = vulnDict(newResults, maxrN)
print('New Counts Returned')

totalnewSeverity = totalSeverity(dnewCounts)
difference = differences(dlastCounts,dnewCounts)

#### write report ####
metricsPath = args.metricsFile
metricsFile = openpyxl.Workbook()
metricsResults = metricsFile.get_active_sheet()

#metrics columns
metricsResults['A1'].value = args.reportDate
metricsResults['B2'].value = 'CVSS < 4'
metricsResults['B3'].value = 'Low'
metricsResults['C2'].value = 'CVSS 4 - 6.9'
metricsResults['C3'].value = 'Medium'
metricsResults['D2'].value = 'CVSS 7-9'
metricsResults['D3'].value = 'High'
metricsResults['E2'].value = 'CVSS 10'
metricsResults['E3'].value = 'Critical'
metricsResults['A4'].value = 'Found'
metricsResults['A5'].value = 'Remediated'
metricsResults['A6'].value = 'Not Remediated'

# Found (row 4)
metricsResults['B4'].value = difference['low']['discovered']
metricsResults['C4'].value = difference['medium']['discovered']
metricsResults['D4'].value = difference['high']['discovered']
metricsResults['E4'].value = difference['critical']['discovered']

# Remediated (row 5)
metricsResults['B5'].value = difference['low']['remediated']
metricsResults['C5'].value = difference['medium']['remediated']
metricsResults['D5'].value = difference['high']['remediated']
metricsResults['E5'].value = difference['critical']['remediated']

# Not Remediated (row 6)
metricsResults['B6'].value = totalnewSeverity['low']
metricsResults['C6'].value = totalnewSeverity['medium']
metricsResults['D6'].value = totalnewSeverity['high']
metricsResults['E6'].value = totalnewSeverity['critical']

metricsFile.save(metricsPath)
